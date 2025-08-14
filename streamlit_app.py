"""
üè• OneMed Fakturabehandling - Streamlit Edition

Profesjonell fakturabehandling med OneMed-farger og lys/m√∏rk modus
Automatisk kontering av Telia Norge AS fakturaer med norsk finansterminologi

Kj√∏r med: streamlit run streamlit_app.py
"""

import streamlit as st
import sys
import os
from pathlib import Path
import tempfile
import json

# Add src to path for imports
sys.path.append('src')

# Configure page
st.set_page_config(
    page_title="OneMed Fakturabehandling", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# OneMed EKTE Professional Healthcare Color Palette - Sleek & Modern
ONEMED_COLORS = {
    "light": {
        "primary": "#497886",          # OneMed FAKTISK bakgrunnsfarge
        "primary_dark": "#3a5f6b",     # M√∏rkere teal
        "secondary": "#6b9bb0",        # Lysere teal
        "accent": "#4CAF50",           # Success Green
        "warning": "#FF9800",          # Warning Amber
        "error": "#F44336",            # Error Red
        "background": "#FAFAFA",       # Ren hvit bakgrunn
        "surface": "#FFFFFF",          # Hvite kort
        "text": "#212121",             # M√∏rk tekst
        "text_secondary": "#757575",   # Gr√• tekst
        "border": "#E0E0E0",           # Lys border
        "shadow": "rgba(73, 120, 134, 0.1)"  # OneMed shadow
    },
    "dark": {
        "primary": "#5a8a98",          # Lysere OneMed for m√∏rk modus
        "primary_dark": "#497886",     # Original OneMed
        "secondary": "#7ba8bc",        # Lysere teal
        "accent": "#66BB6A",           # Lysere gr√∏nn
        "warning": "#FFA726",          # Lysere amber
        "error": "#EF5350",            # Lysere r√∏d
        "background": "#121212",       # Material m√∏rk
        "surface": "#1E1E1E",          # M√∏rke kort
        "text": "#FFFFFF",             # Hvit tekst
        "text_secondary": "#AAAAAA",  # Lys gr√• tekst
        "border": "#333333",           # M√∏rk border
        "shadow": "rgba(0, 0, 0, 0.3)" # M√∏rk shadow
    }
}

# Initialize theme in session state
if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = False

def apply_onemed_theme():
    """Apply SLEEK OneMed Professional Design - Enterprise Level"""
    theme = "dark" if st.session_state.dark_mode else "light"
    colors = ONEMED_COLORS[theme]
    
    css = f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
        
        /* ENTERPRISE LEVEL STYLING - OneMed Professional */
        .stApp {{
            background: {colors["background"]};
            color: {colors["text"]};
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
            line-height: 1.6;
        }}
        
        /* OneMed SLEEK Header with Corporate Styling */
        .onemed-header {{
            background: linear-gradient(135deg, {colors["primary"]} 0%, {colors["primary_dark"]} 100%);
            padding: 3rem 2rem;
            border-radius: 0;
            margin: -1rem -2rem 2rem -2rem;
            position: relative;
            box-shadow: 0 8px 32px {colors["shadow"]};
            border-bottom: 3px solid {colors["primary_dark"]};
        }}
        
        .onemed-header::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: url('data:image/svg+xml,<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100"><defs><pattern id="grid" width="10" height="10" patternUnits="userSpaceOnUse"><path d="M 10 0 L 0 0 0 10" fill="none" stroke="rgba(255,255,255,0.1)" stroke-width="0.5"/></pattern></defs><rect width="100" height="100" fill="url(%23grid)" /></svg>');
            opacity: 0.1;
        }}
        
        .onemed-logo {{
            display: flex;
            align-items: center;
            gap: 1rem;
            margin-bottom: 1rem;
        }}
        
        .onemed-logo h1 {{
            color: white;
            margin: 0;
            font-size: 2.8rem;
            font-weight: 700;
            letter-spacing: -0.02em;
            text-shadow: 0 2px 4px rgba(0,0,0,0.2);
        }}
        
        .onemed-logo-icon {{
            width: 60px;
            height: 60px;
            background: rgba(255,255,255,0.15);
            border-radius: 12px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 2rem;
            backdrop-filter: blur(10px);
        }}
        
        /* Professional Sidebar */
        .css-1d391kg, .css-1cypcdb {{
            background: {colors["surface"]};
            border-right: 1px solid {colors["border"]};
            backdrop-filter: blur(20px);
        }}
        
        /* Enterprise Metrics Cards */
        [data-testid="metric-container"] {{
            background: {colors["surface"]};
            border: 1px solid {colors["border"]};
            border-radius: 16px;
            padding: 1.5rem;
            box-shadow: 0 4px 20px {colors["shadow"]};
            transition: all 0.3s ease;
            position: relative;
            overflow: hidden;
        }}
        
        [data-testid="metric-container"]:hover {{
            transform: translateY(-2px);
            box-shadow: 0 8px 32px {colors["shadow"]};
        }}
        
        [data-testid="metric-container"]::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 4px;
            background: linear-gradient(90deg, {colors["primary"]}, {colors["secondary"]});
        }}
        
        /* Premium File Uploader */
        [data-testid="stFileUploader"] {{
            background: {colors["surface"]};
            border: 2px dashed {colors["border"]};
            border-radius: 20px;
            padding: 3rem 2rem;
            text-align: center;
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            overflow: hidden;
        }}
        
        [data-testid="stFileUploader"]::before {{
            content: '';
            position: absolute;
            top: -2px;
            left: -2px;
            right: -2px;
            bottom: -2px;
            background: linear-gradient(45deg, {colors["primary"]}, {colors["secondary"]});
            border-radius: 20px;
            opacity: 0;
            transition: opacity 0.4s ease;
            z-index: -1;
        }}
        
        [data-testid="stFileUploader"]:hover {{
            border-color: transparent;
            transform: translateY(-4px);
            box-shadow: 0 12px 40px {colors["shadow"]};
        }}
        
        [data-testid="stFileUploader"]:hover::before {{
            opacity: 1;
        }}
        
        /* Executive Level Buttons */
        .stButton > button {{
            background: linear-gradient(135deg, {colors["primary"]} 0%, {colors["primary_dark"]} 100%);
            color: white;
            border: none;
            border-radius: 12px;
            font-weight: 600;
            font-size: 1rem;
            padding: 1rem 2rem;
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            overflow: hidden;
            box-shadow: 0 4px 16px {colors["shadow"]};
            letter-spacing: 0.02em;
        }}
        
        .stButton > button:hover {{
            transform: translateY(-3px);
            box-shadow: 0 8px 32px {colors["shadow"]};
        }}
        
        .stButton > button:active {{
            transform: translateY(-1px);
        }}
        
        /* Corporate Messages */
        .stSuccess {{
            background: linear-gradient(135deg, rgba(76, 175, 80, 0.1), rgba(76, 175, 80, 0.05));
            border: 1px solid {colors["accent"]};
            border-left: 4px solid {colors["accent"]};
            border-radius: 12px;
            backdrop-filter: blur(20px);
        }}
        
        .stError {{
            background: linear-gradient(135deg, rgba(244, 67, 54, 0.1), rgba(244, 67, 54, 0.05));
            border: 1px solid {colors["error"]};
            border-left: 4px solid {colors["error"]};
            border-radius: 12px;
            backdrop-filter: blur(20px);
        }}
        
        .stWarning {{
            background: linear-gradient(135deg, rgba(255, 152, 0, 0.1), rgba(255, 152, 0, 0.05));
            border: 1px solid {colors["warning"]};
            border-left: 4px solid {colors["warning"]};
            border-radius: 12px;
            backdrop-filter: blur(20px);
        }}
        
        .stInfo {{
            background: linear-gradient(135deg, rgba(73, 120, 134, 0.1), rgba(73, 120, 134, 0.05));
            border: 1px solid {colors["primary"]};
            border-left: 4px solid {colors["primary"]};
            border-radius: 12px;
            backdrop-filter: blur(20px);
        }}
        
        /* Professional Expanders */
        .streamlit-expanderHeader {{
            background: {colors["surface"]};
            border-radius: 12px;
            border: 1px solid {colors["border"]};
            box-shadow: 0 2px 8px {colors["shadow"]};
            transition: all 0.3s ease;
        }}
        
        .streamlit-expanderHeader:hover {{
            box-shadow: 0 4px 16px {colors["shadow"]};
        }}
        
        /* Executive Containers */
        .onemed-container {{
            background: {colors["surface"]};
            border: 1px solid {colors["border"]};
            border-radius: 16px;
            padding: 2rem;
            margin-bottom: 1.5rem;
            box-shadow: 0 4px 20px {colors["shadow"]};
            backdrop-filter: blur(20px);
            position: relative;
            overflow: hidden;
        }}
        
        .onemed-container::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 3px;
            background: linear-gradient(90deg, {colors["primary"]}, {colors["secondary"]});
        }}
        
        /* Premium Text Styling */
        .text-primary {{ 
            color: {colors["primary"]}; 
            font-weight: 600;
        }}
        .text-secondary {{ 
            color: {colors["text_secondary"]}; 
            font-weight: 400;
        }}
        .text-success {{ 
            color: {colors["accent"]}; 
            font-weight: 600;
        }}
        .text-error {{ 
            color: {colors["error"]}; 
            font-weight: 600;
        }}
        .text-warning {{ 
            color: {colors["warning"]}; 
            font-weight: 600;
        }}
        
        /* Executive DataFrames */
        .dataframe {{
            border-radius: 12px;
            border: 1px solid {colors["border"]};
            box-shadow: 0 4px 16px {colors["shadow"]};
            overflow: hidden;
        }}
        
        /* Professional Typography */
        h1, h2, h3, h4, h5, h6 {{
            color: {colors["text"]};
            font-weight: 600;
            letter-spacing: -0.01em;
        }}
        
        /* Hide Streamlit Branding */
        #MainMenu {{visibility: hidden;}}
        footer {{visibility: hidden;}}
        header {{visibility: hidden;}}
        
        /* Premium Scrollbar */
        ::-webkit-scrollbar {{
            width: 12px;
        }}
        
        ::-webkit-scrollbar-track {{
            background: {colors["background"]};
            border-radius: 6px;
        }}
        
        ::-webkit-scrollbar-thumb {{
            background: linear-gradient(135deg, {colors["primary"]}, {colors["secondary"]});
            border-radius: 6px;
            box-shadow: 0 2px 4px {colors["shadow"]};
        }}
        
        ::-webkit-scrollbar-thumb:hover {{
            background: linear-gradient(135deg, {colors["primary_dark"]}, {colors["primary"]});
        }}
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)

# Initialize parser (with fallback if not available)
@st.cache_resource
def get_parser():
    try:
        from extraction.suppliers.telia import TeliaParser
        return TeliaParser()
    except ImportError:
        return None

def extract_text_fallback(pdf_file):
    """Fallback PDF extraction if main parser not available"""
    try:
        import PyPDF2
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except ImportError:
        return None

def main():
    # Apply OneMed theme
    apply_onemed_theme()
    
    # Sidebar with theme toggle and controls
    with st.sidebar:
        st.markdown("### OneMed Innstillinger")
        
        # Theme toggle
        theme_text = "Bytt til lys modus" if st.session_state.dark_mode else "Bytt til m√∏rk modus"
        
        if st.button(theme_text, key="theme_toggle", use_container_width=True):
            st.session_state.dark_mode = not st.session_state.dark_mode
            st.rerun()
        
        st.markdown("---")
        
        # OneMed branding info
        st.markdown("""
        <div class="onemed-container">
            <h4 class="text-primary">OneMed</h4>
            <p class="text-secondary" style="font-size: 0.9rem;">
                Profesjonell fakturabehandling for helseteknologi
            </p>
            <p class="text-secondary" style="font-size: 0.8rem;">
                Norsk finansterminologi
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # System info
        with st.expander("Systeminfo"):
            current_theme = "M√∏rk modus" if st.session_state.dark_mode else "Lys modus"
            st.write(f"**Tema:** {current_theme}")
            st.write(f"**Versjon:** 1.0")
            st.write(f"**Leverand√∏r:** Telia Norge AS")
            
            parser = get_parser()
            if parser:
                st.write("**Status:** Full funksjonalitet")
            else:
                st.write("**Status:** Begrenset modus")
    
    # Professional OneMed Header
    st.markdown("""
    <div class="onemed-header">
        <div class="onemed-logo">
            <div class="onemed-logo-icon">OM</div>
            <div>
                <h1>OneMed Fakturabehandling</h1>
                <p style="color: rgba(255,255,255,0.85); margin: 0; font-size: 1.2rem; font-weight: 300; letter-spacing: 0.02em;">
                    Automatisk kostnadsallokering og finansrapportering for Telia Norge AS fakturaer
                </p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Instructions
    with st.expander("Slik bruker du OneMed Fakturabehandling", expanded=True):
        st.markdown("""
        <div class="onemed-container">
            <ol style="font-size: 1rem; line-height: 1.8;">
                <li><span class="text-primary"><strong>Last opp</strong></span> en Telia Norge AS faktura (PDF-format)</li>
                <li><span class="text-secondary"><strong>Last opp</strong></span> kostnadsb√¶rer-fil (Excel) - valgfritt</li>
                <li><span class="text-success"><strong>Klikk</strong></span> "Behandle Faktura" for automatisk kostnadsallokering</li>
                <li><span class="text-warning"><strong>Gjennomg√•</strong></span> resultater og generer finansrapport</li>
            </ol>
            <div style="margin-top: 1rem; padding: 1rem; background-color: rgba(73, 120, 134, 0.1); border-radius: 8px;">
                <h5 class="text-primary">Norsk Finansterminologi:</h5>
                <ul style="font-size: 0.9rem;">
                    <li><strong>ALLOKERT</strong> - Kostnad automatisk tildelt riktig kostsenter</li>
                    <li><strong>KREVER_MANUELL_BEHANDLING</strong> - Kostnad som trenger manuell vurdering</li>
                    <li><strong>FLERE_MULIGE_KOSTSENTRE</strong> - Tvetydige ansattnavn som krever avklaring</li>
                </ul>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("### Filopplasting")
    
    # File uploads
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        <div class="onemed-container">
            <h4 class="text-primary">Telia Faktura (PDF)</h4>
            <p class="text-secondary">P√•krevd - Telia Norge AS fakturaer</p>
        </div>
        """, unsafe_allow_html=True)
        
        pdf_file = st.file_uploader(
            "Velg PDF-fil", 
            type=['pdf'], 
            help="Last opp Telia Norge AS faktura i PDF-format",
            key="pdf_upload"
        )
        
        if pdf_file:
            st.success(f"PDF lastet opp: {pdf_file.name} ({pdf_file.size/1024:.1f} KB)")
    
    with col2:
        st.markdown("""
        <div class="onemed-container">
            <h4 class="text-secondary">Kostnadsb√¶rere (Excel)</h4>
            <p class="text-secondary">Valgfritt - bruker mock-data hvis tom</p>
        </div>
        """, unsafe_allow_html=True)
        
        excel_file = st.file_uploader(
            "Velg Excel-fil", 
            type=['xlsx', 'xls'], 
            help="Valgfritt - bruker mock-data hvis ikke angitt",
            key="excel_upload"
        )
        
        if excel_file:
            st.success(f"Excel lastet opp: {excel_file.name} ({excel_file.size/1024:.1f} KB)")
        else:
            st.info("Bruker OneMed mock kostnadsb√¶rer-data")
    
    # Process button
    st.markdown("---")
    col_center = st.columns([1, 2, 1])[1]
    with col_center:
        if pdf_file:
            if st.button("Behandle Faktura", type="primary", use_container_width=True):
                process_invoice(pdf_file, excel_file)
        else:
            st.button("Last opp PDF-fil f√∏rst", disabled=True, use_container_width=True)

def process_invoice(pdf_file, excel_file):
    """Process the uploaded invoice with professional OneMed analysis"""
    
    parser = get_parser()
    
    with st.spinner("Behandler faktura med OneMed Fakturabehandling..."):
        try:
            # Save PDF to temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                tmp_file.write(pdf_file.getvalue())
                pdf_path = tmp_file.name
            
            # Extract text
            pdf_content = None
            if parser:
                try:
                    from extraction.suppliers.telia import extract_text_from_pdf
                    pdf_content = extract_text_from_pdf(Path(pdf_path))
                except:
                    pass
            
            # Fallback extraction
            if not pdf_content:
                with open(pdf_path, 'rb') as f:
                    pdf_content = extract_text_fallback(f)
            
            # Clean up temp file
            os.unlink(pdf_path)
            
            if not pdf_content:
                st.error("Kunne ikke lese tekst fra PDF-filen")
                return
            
            # Check if it's a Telia invoice (basic check)
            if "Telia" not in pdf_content and "telia" not in pdf_content.lower():
                st.error("Dette ser ikke ut til √• v√¶re en Telia Norge AS faktura")
                return
            
            # ADVANCED DEBUG: Show what's happening with the PDF
            with st.expander("OneMed DEBUG - Se hva som skjer med PDF-en", expanded=False):
                st.markdown("### PDF Analyse")
                col1, col2 = st.columns(2)
                
                with col1:
                    st.metric("PDF Tekst lengde", f"{len(pdf_content):,} tegn")
                    contains_telia = "Telia" in pdf_content or "telia" in pdf_content.lower()
                    st.metric("Inneholder 'Telia'", "JA" if contains_telia else "NEI")
                
                with col2:
                    parser_available = parser is not None
                    st.metric("TeliaParser tilgjengelig", "JA" if parser_available else "NEI")
                    if parser_available:
                        can_parse = parser.can_parse(pdf_content)
                        st.metric("kan_parse()", "JA" if can_parse else "NEI")
                    
                if parser_available:
                    st.markdown("### Pattern Matching Test")
                    patterns = parser.get_identification_patterns()
                    import re
                    matches = 0
                    for i, pattern in enumerate(patterns, 1):
                        found = re.search(pattern, pdf_content, re.IGNORECASE)
                        status = "FUNNET" if found else "IKKE FUNNET"
                        st.write(f"{i}. `{pattern}` ‚Üí {status}")
                        if found:
                            matches += 1
                    st.write(f"**Totalt: {matches}/5 patterns funnet** (trenger 3+ for parsing)")
                
                st.markdown("### PDF Tekst Sample")
                st.text_area("F√∏rste 500 tegn fra PDF:", pdf_content[:500], height=200)

            # Process with parser or create mock data
            if parser and parser.can_parse(pdf_content):
                st.success("OneMed TeliaParser kan parse denne PDF-en!")
                result = parser.parse_invoice_with_cost_bearers(pdf_content)
                display_results(result, is_full_parser=True)
            else:
                if parser:
                    st.warning("PDF-en passerte ikke TeliaParser validering - bruker demo-data")
                    if not parser.can_parse(pdf_content):
                        st.error("**√Örsak**: Mindre enn 3 av 5 identifikasjonsm√∏nstre ble funnet")
                        st.info("**Tips**: Sjekk at PDF-en inneholder 'Fakturanummer:', 'Telia Norge AS', og 'SUM DENNE PERIODE'")
                else:
                    st.warning("TeliaParser ikke tilgjengelig - bruker demo-data")
                
                # Create mock processing results for demo
                mock_result = create_mock_result(pdf_content, pdf_file.name)
                display_results(mock_result, is_full_parser=False)
                
        except Exception as e:
            st.error(f"Feil ved behandling av faktura: {str(e)}")

def create_mock_result(pdf_content, filename):
    """Create mock processing results for demo purposes"""
    return {
        'leverandor': {'navn': 'Telia Norge AS'},
        'faktura_metadata': {
            'fakturanummer': '123456', 
            'fakturadato': '15.01.2024',
            'periode_fra': '01.01.2024',
            'periode_til': '31.01.2024'
        },
        'bel√∏p_sammendrag': {
            'totalbel√∏p': 1426.50,
            'valuta': 'NOK'
        },
        'kostnadsbarer_telia': [
            {
                'navn_fra_faktura': 'Annlaug Amundsen',
                'matched_fornavn': 'Annlaug',
                'matched_etternavn': 'Amundsen',
                'kostsenter': 1001,
                'telefonnummer': '91854560',
                'sum_denne_periode': 450.75,
                'match_status': 'ALLOKERT',
                'confidence_score': 0.95,
                'deviation_reason': ''
            },
            {
                'navn_fra_faktura': 'Andreas Hansen',
                'matched_fornavn': 'Andreas',
                'matched_etternavn': 'Hansen',
                'kostsenter': 1002,
                'telefonnummer': '92078335',
                'sum_denne_periode': 320.50,
                'match_status': 'ALLOKERT',
                'confidence_score': 0.90,
                'deviation_reason': ''
            },
            {
                'navn_fra_faktura': 'Allan Simonsen',
                'matched_fornavn': 'Allan',
                'matched_etternavn': 'Simonsen',
                'kostsenter': 1003,
                'telefonnummer': '90063358',
                'sum_denne_periode': 275.00,
                'match_status': 'ALLOKERT',
                'confidence_score': 0.88,
                'deviation_reason': ''
            },
            {
                'navn_fra_faktura': 'Dr. Maria Lindstr√∂m',
                'matched_fornavn': '',
                'matched_etternavn': '',
                'kostsenter': None,
                'telefonnummer': '90123456',
                'sum_denne_periode': 380.25,
                'match_status': 'KREVER_MANUELL_BEHANDLING',
                'confidence_score': 0.0,
                'deviation_reason': 'Ansatt "Maria Lindstr√∂m" ikke funnet i OneMed kostnadsb√¶rer-registeret'
            }
        ],
        'linjedetaljer': [
            {'produktnavn': 'Tjeneste for Annlaug Amundsen', 'employee_name': 'Annlaug Amundsen', 'phone_number': '91854560', 'linjesum': 450.75},
            {'produktnavn': 'Tjeneste for Andreas Hansen', 'employee_name': 'Andreas Hansen', 'phone_number': '92078335', 'linjesum': 320.50},
            {'produktnavn': 'Tjeneste for Allan Simonsen', 'employee_name': 'Allan Simonsen', 'phone_number': '90063358', 'linjesum': 275.00},
            {'produktnavn': 'Tjeneste for Maria Lindstr√∂m', 'employee_name': 'Maria Lindstr√∂m', 'phone_number': '90123456', 'linjesum': 380.25}
        ],
        'kvalitetskontroll': {
            'unmatched_count': 1,
            'processing_confidence': 0.93,
            'requires_manual_review': True,
            'validation_errors': ['1 kostnad krever manuell behandling'],
            'totalbel√∏p_stemmer': True
        }
    }

def display_results(invoice_data, is_full_parser=True):
    """Display processing results with professional OneMed styling and Norwegian financial terminology"""
    
    # Success header
    parser_status = "Full OneMed Parser" if is_full_parser else "Demo Modus"
    st.markdown(f"""
    <div class="onemed-container" style="text-align: center; border-left: 4px solid #4CAF50;">
        <h2 class="text-success">Faktura behandlet med OneMed Fakturabehandling</h2>
        <p class="text-secondary">Automatisk kostnadsallokering fullf√∏rt med norsk finansterminologi ({parser_status})</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Summary metrics  
    st.markdown("## OneMed Kostnadsallokering - Sammendrag")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "Totalt Bel√∏p",
            f"{invoice_data['bel√∏p_sammendrag']['totalbel√∏p']:.2f} NOK",
            help="Totalt fakturabel√∏p i norske kroner"
        )
    
    with col2:
        allokerte = len([cb for cb in invoice_data['kostnadsbarer_telia'] 
                       if cb['match_status'] == 'ALLOKERT'])
        totalt = len(invoice_data['kostnadsbarer_telia'])
        delta = f"+{allokerte}" if allokerte > 0 else None
        st.metric("Allokerte Kostnader", f"{allokerte}/{totalt}", delta=delta)
    
    with col3:
        avvik = invoice_data['kvalitetskontroll']['unmatched_count']
        delta_color = "inverse" if avvik > 0 else "normal"
        st.metric("Manuelle Behandlinger", avvik, 
                 delta=f"{avvik} krever manuell behandling" if avvik > 0 else "Ingen avvik",
                 delta_color=delta_color)
    
    with col4:
        tillit = invoice_data['kvalitetskontroll']['processing_confidence'] * 100
        delta_color = "normal" if tillit >= 90 else "inverse"
        st.metric("Systemtillit", f"{tillit:.0f}%", 
                 delta="H√∏y tillit" if tillit >= 90 else "Moderat tillit",
                 delta_color=delta_color)
    
    # Quality control alerts
    if invoice_data['kvalitetskontroll']['validation_errors']:
        st.markdown("## OneMed Kvalitetskontroll")
        st.markdown("""
        <div class="onemed-container" style="border-left: 4px solid #FF9800;">
            <h4 class="text-warning">Viktige merknader for OneMed</h4>
            <p class="text-secondary">F√∏lgende punkter krever oppmerksomhet fra finansavdelingen:</p>
        </div>
        """, unsafe_allow_html=True)
        
        for error in invoice_data['kvalitetskontroll']['validation_errors']:
            st.warning(f"{error}")
    
    # Cost allocation results
    st.markdown("## OneMed Kostnadsallokering")
    st.markdown("""
    <div class="onemed-container">
        <p class="text-secondary">Resultater fra automatisk kostnadsallokering mot OneMed kostnadsb√¶rer-registeret</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Create structured table for cost allocation
    for i, cb in enumerate(invoice_data['kostnadsbarer_telia']):
        with st.container():
            cols = st.columns([3, 1, 1, 2])
            
            with cols[0]:
                if cb['match_status'] == 'ALLOKERT':
                    st.success(f"**{cb['navn_fra_faktura']}**")
                    if cb['matched_fornavn'] and cb['matched_etternavn']:
                        st.caption(f"Matchet: {cb['matched_fornavn']} {cb['matched_etternavn']}")
                else:
                    st.error(f"**{cb['navn_fra_faktura']}**")
                    if cb['deviation_reason']:
                        st.caption(f"√Örsak: {cb['deviation_reason']}")
            
            with cols[1]:
                if cb['kostsenter']:
                    st.markdown(f"**Kostsenter:**<br>{cb['kostsenter']}", unsafe_allow_html=True)
                else:
                    st.markdown("**Kostsenter:**<br>‚Äì", unsafe_allow_html=True)
            
            with cols[2]:
                status_display = 'ALLOKERT' if cb['match_status'] == 'ALLOKERT' else 'MANUELL'
                st.markdown(f"**Status:**<br>{status_display}", unsafe_allow_html=True)
            
            with cols[3]:
                st.markdown(f"**Bel√∏p:**<br>{cb['sum_denne_periode']:.2f} NOK", 
                           unsafe_allow_html=True)
        
        st.divider()
    
    # Invoice details
    with st.expander("OneMed Fakturaopplysninger", expanded=False):
        st.markdown("""
        <div class="onemed-container">
            <h4 class="text-primary">Uttrukket informasjon fra Telia Norge AS faktura</h4>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown(f"""
            **Leverand√∏r:** {invoice_data['leverandor']['navn']}  
            **Fakturanummer:** {invoice_data['faktura_metadata']['fakturanummer']}  
            **Fakturadato:** {invoice_data['faktura_metadata']['fakturadato']}
            """)
        
        with col2:
            if invoice_data['faktura_metadata'].get('periode_fra'):
                st.markdown(f"""
                **Periode:** {invoice_data['faktura_metadata']['periode_fra']} - {invoice_data['faktura_metadata']['periode_til']}  
                **Totalbel√∏p:** {invoice_data['bel√∏p_sammendrag']['totalbel√∏p']:.2f} {invoice_data['bel√∏p_sammendrag']['valuta']}  
                **Kvalitet:** {invoice_data['kvalitetskontroll']['processing_confidence']*100:.1f}% systemtillit
                """)
    
    # Line details table
    with st.expander("OneMed Linjedetaljer", expanded=False):
        st.markdown("""
        <div class="onemed-container">
            <h4 class="text-primary">Individuelle tjenester fra fakturaen</h4>
        </div>
        """, unsafe_allow_html=True)
        
        import pandas as pd
        
        lines_data = []
        for line in invoice_data['linjedetaljer']:
            lines_data.append({
                'Tjeneste': line['produktnavn'],
                'Ansatt': line['employee_name'],
                'Telefon': line['phone_number'],
                'Bel√∏p (NOK)': f"{line['linjesum']:.2f}"
            })
        
        df = pd.DataFrame(lines_data)
        st.dataframe(df, use_container_width=True, hide_index=True)
    
    # Finansrapport generering  
    st.markdown("---")
    st.markdown("## Finansrapport for OneMed")
    
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col1:
        if st.button("Behandle Ny Faktura", use_container_width=True):
            st.rerun()
    
    with col2:
        # Create comprehensive finance report
        finance_report = create_finance_report(invoice_data)
        st.download_button(
            label="Last ned Finansrapport (Excel)",
            data=finance_report['excel_data'],
            file_name=f"onemed_kostnadsrapport_{invoice_data['faktura_metadata']['fakturanummer']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    with col3:
        # Create CSV export for accounting systems
        csv_data = create_accounting_csv(invoice_data)
        st.download_button(
            label="Last ned CSV (Regnskapssystem)",
            data=csv_data,
            file_name=f"onemed_kontering_{invoice_data['faktura_metadata']['fakturanummer']}.csv",
            mime="text/csv",
            use_container_width=True
        )

def create_finance_report(invoice_data):
    """Create a comprehensive Excel report for finance department"""
    import io
    from datetime import datetime
    
    # Create in-memory Excel file
    excel_buffer = io.BytesIO()
    
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Sammendrag"
        
        # Header
        ws_summary['A1'] = "OneMed Kostnadsrapport - Telia Norge AS"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A3'] = f"Fakturanummer: {invoice_data['faktura_metadata']['fakturanummer']}"
        ws_summary['A4'] = f"Fakturadato: {invoice_data['faktura_metadata']['fakturadato']}"
        ws_summary['A5'] = f"Totalbel√∏p: {invoice_data['bel√∏p_sammendrag']['totalbel√∏p']:.2f} NOK"
        ws_summary['A6'] = f"Generert: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        # Cost allocation summary
        ws_summary['A8'] = "Kostnadsallokering:"
        ws_summary['A8'].font = Font(bold=True)
        
        allokerte = len([cb for cb in invoice_data['kostnadsbarer_telia'] if cb['match_status'] == 'ALLOKERT'])
        manuelle = invoice_data['kvalitetskontroll']['unmatched_count']
        
        ws_summary['A9'] = f"Automatisk allokerte kostnader: {allokerte}"
        ws_summary['A10'] = f"Krever manuell behandling: {manuelle}"
        ws_summary['A11'] = f"Systemtillit: {invoice_data['kvalitetskontroll']['processing_confidence']*100:.1f}%"
        
        # Cost allocation details sheet
        ws_details = wb.create_sheet("Kostnadsfordeling")
        
        # Headers for cost allocation
        headers = ['Ansatt', 'Fornavn', 'Etternavn', 'Kostsenter', 'Telefon', 'Bel√∏p (NOK)', 'Status', 'Avviksbegrunnelse']
        for idx, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="497886", end_color="497886", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data rows
        for row_idx, cb in enumerate(invoice_data['kostnadsbarer_telia'], 2):
            ws_details.cell(row=row_idx, column=1).value = cb['navn_fra_faktura']
            ws_details.cell(row=row_idx, column=2).value = cb['matched_fornavn']
            ws_details.cell(row=row_idx, column=3).value = cb['matched_etternavn']
            ws_details.cell(row=row_idx, column=4).value = cb['kostsenter']
            ws_details.cell(row=row_idx, column=5).value = cb['telefonnummer']
            ws_details.cell(row=row_idx, column=6).value = cb['sum_denne_periode']
            ws_details.cell(row=row_idx, column=7).value = cb['match_status']
            ws_details.cell(row=row_idx, column=8).value = cb['deviation_reason']
            
            # Color code based on status
            if cb['match_status'] == 'ALLOKERT':
                fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            for col in range(1, 9):
                ws_details.cell(row=row_idx, column=col).fill = fill
        
        # Auto-adjust column widths
        for column in ws_details.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws_details.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return {
            'excel_data': excel_buffer.getvalue(),
            'filename': f"onemed_kostnadsrapport_{invoice_data['faktura_metadata']['fakturanummer']}.xlsx"
        }
        
    except ImportError:
        # Fallback to simple JSON if Excel libraries not available
        import json
        json_data = {
            'onemed_finansrapport': {
                'faktura_metadata': invoice_data['faktura_metadata'],
                'kostnadsbarer': invoice_data['kostnadsbarer_telia'],
                'kvalitetskontroll': invoice_data['kvalitetskontroll'],
                'sammendrag': {
                    'totalt_bel√∏p': invoice_data['bel√∏p_sammendrag']['totalbel√∏p'],
                    'allokerte_kostnader': len([cb for cb in invoice_data['kostnadsbarer_telia'] if cb['match_status'] == 'ALLOKERT']),
                    'manuelle_behandlinger': invoice_data['kvalitetskontroll']['unmatched_count']
                }
            }
        }
        return {
            'excel_data': json.dumps(json_data, indent=2, ensure_ascii=False).encode('utf-8'),
            'filename': f"onemed_finansrapport_{invoice_data['faktura_metadata']['fakturanummer']}.json"
        }

def create_accounting_csv(invoice_data):
    """Create CSV export formatted for Norwegian accounting systems"""
    import io
    from datetime import datetime
    
    csv_buffer = io.StringIO()
    
    # Write CSV header
    csv_buffer.write("Fakturanummer,Ansatt,Kostsenter,Bel√∏p,Valuta,Status,Dato,Bilag\n")
    
    # Write data rows
    for cb in invoice_data['kostnadsbarer_telia']:
        row = [
            invoice_data['faktura_metadata']['fakturanummer'],
            cb['navn_fra_faktura'],
            cb['kostsenter'] or 'UALLOKERT',
            f"{cb['sum_denne_periode']:.2f}",
            'NOK',
            cb['match_status'],
            invoice_data['faktura_metadata']['fakturadato'],
            f"TELIA-{invoice_data['faktura_metadata']['fakturanummer']}"
        ]
        csv_buffer.write(','.join(map(str, row)) + '\n')
    
    csv_content = csv_buffer.getvalue()
    csv_buffer.close()
    
    return csv_content.encode('utf-8')

if __name__ == "__main__":
    main()